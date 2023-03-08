import openpyxl as xl
import sqlite3
from openpyxl.chart import BarChart, Reference
import pandas as pd
import argparse
from configparser import ConfigParser
import os

con = sqlite3.connect("../db.sqlite3")


def etl_load_data(filename):
    df = pd.read_excel(filename)
    try:
        df.to_sql("etl_loader_customer", con, if_exists="append", index=False)
    except Exception as e:
        print("FAILURE TO APPEND: {}".format(e))


def etl_load_workbook(filename, graph_filename):
    wb = xl.load_workbook(filename)
    sheet = wb["DATA"]
    for row in range(2, sheet.max_row + 1):
        row_id = sheet.cell(row, 1).value
        first_name = sheet.cell(row, 2).value
        last_name = sheet.cell(row, 3).value
        email = sheet.cell(row, 4).value
        gender = sheet.cell(row, 5).value
        ip_address = sheet.cell(row, 6).value

        print(row_id, first_name, last_name, email, gender, ip_address)

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "h2")
    #wb.save(graph_filename)  # Work in progress


def main():

    parser = argparse.ArgumentParser(
        description="Excel File Name expected, ")
    parser.add_argument('--file', help='foo help')

    args = parser.parse_args()

    if args.file:

        try:
            csv_file = (args.file)
            graph_file = '../csv_io/graph.xlsx'
        except Exception as e:
            print("INVALID FILE: {}".format(e))
    else:

        # Read config.ini file
        config_object = ConfigParser()
        config_object.read("config.ini")

        # Get the csv path
        csv_section = config_object["CSVFILEPATH"]
        print(csv_section)
        csv_file = csv_section["CSV_FILE"]
        graph_file = csv_section["GRAPH_FILE"]


    etl_load_workbook(csv_file, graph_file)
    etl_load_data(csv_file)


if __name__ == "__main__":
    main()
